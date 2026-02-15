# components/utils/excel_generator.py
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import io
from typing import Dict, Union, List, Optional
import os

# デフォルトのテンプレートファイルパス（配置場所に合わせて変更してください）
DEFAULT_TEMPLATE_PATH = "■初回送付セット【20251218版】　.xlsx"

def fill_initial_set_excel(
    json_data: Dict[str, str], 
    template_file: Optional[Union[str, io.BytesIO]] = None
) -> io.BytesIO:
    """
    KintoneのJSONデータを基に、初回送付セットExcelの「基本情報入力」シートに値を転記します。

    Args:
        json_data (Dict[str, str]): Kintoneから取得したJSONデータ（辞書型）
        template_file (Optional[Union[str, io.BytesIO]]): テンプレートExcelファイル。
            指定がない場合はデフォルトパスを使用。

    Returns:
        io.BytesIO: 編集後のExcelバイナリデータ（ダウンロード用）
    
    Raises:
        FileNotFoundError: テンプレートファイルが見つからない場合
        KeyError: 指定されたシートが存在しない場合
    """
    
    # テンプレートの読み込み元を決定
    source = template_file if template_file else DEFAULT_TEMPLATE_PATH
    
    if isinstance(source, str) and not os.path.exists(source):
        raise FileNotFoundError(f"テンプレートファイルが見つかりません: {source}")

    # Excelブックを開く
    wb = openpyxl.load_workbook(source)
    
    target_sheet_name = "基本情報入力"
    if target_sheet_name not in wb.sheetnames:
        raise KeyError(f"テンプレート内に '{target_sheet_name}' シートが見つかりません。")
    
    ws: Worksheet = wb[target_sheet_name]

    # マッピング定義
    # JSONのキー : Excelのセル番地（単一文字列 または 文字列のリスト）
    mapping: Dict[str, Union[str, List[str]]] = {
        "顧客コード_2": "B9",
        "顧客名": ["B10", "C24"],  # 複数セルへの転記
        "◎提案項目": "B11",
        "拠点": "B12",
        "担当者①": "B13",
        "担当者②": "D13",
        "被相続人名": "C23",
        "被相続人名（ふりがな）": "D23",
        "相続開始日": "F23",
        "顧客名(ふりがな)": "D24",
        "郵便番号": "G24",
        "住所": "H24",
        "TEL": "J24"
    }

    # データの転記処理
    for json_key, cell_target in mapping.items():
        # JSONから値を取得（キーがない場合は空文字）
        value = json_data.get(json_key, "")
        
        # 転記実行
        if isinstance(cell_target, list):
            for cell_address in cell_target:
                ws[cell_address].value = value
        else:
            ws[cell_target].value = value

    # メモリ上のバイナリとして保存
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    
    return output_buffer